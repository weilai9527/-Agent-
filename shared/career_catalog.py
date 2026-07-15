from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timezone
from difflib import SequenceMatcher
import hashlib
import io
import json
import re
from typing import Any
import unicodedata
from uuid import uuid4


CATALOG_LEVELS = (
    "colleges",
    "majors",
    "training_directions",
    "employment_directions",
    "job_roles",
)
VERSION_STATUSES = {"draft", "published", "archived"}
IMPORT_MODES = {"replace", "merge"}
CATALOG_ENTITY_SPECS = {
    "colleges": {
        "table": "catalog_colleges",
        "parent_field": None,
        "parent_table": None,
        "fields": ("code", "name", "description", "sort_order", "enabled"),
    },
    "majors": {
        "table": "catalog_majors",
        "parent_field": "college_id",
        "parent_table": "catalog_colleges",
        "fields": ("college_id", "code", "name", "degree_type", "description", "sort_order", "enabled"),
    },
    "training_directions": {
        "table": "catalog_training_directions",
        "parent_field": "major_id",
        "parent_table": "catalog_majors",
        "fields": ("major_id", "code", "name", "description", "sort_order", "enabled"),
    },
    "employment_directions": {
        "table": "catalog_employment_directions",
        "parent_field": "training_direction_id",
        "parent_table": "catalog_training_directions",
        "fields": ("training_direction_id", "code", "name", "description", "sort_order", "enabled"),
    },
    "job_roles": {
        "table": "catalog_job_roles",
        "parent_field": "employment_direction_id",
        "parent_table": "catalog_employment_directions",
        "fields": ("employment_direction_id", "code", "name", "description", "sort_order", "enabled"),
    },
    "competencies": {
        "table": "catalog_competencies",
        "parent_field": None,
        "parent_table": None,
        "fields": (
            "code",
            "name",
            "category",
            "description",
            "measurement_hint",
            "sort_order",
            "enabled",
        ),
    },
}


def _id() -> str:
    return str(uuid4())


def _close(cursor: Any) -> None:
    if cursor is not None:
        cursor.close()


def _execute(db: Any, sql: str, params: tuple = ()) -> None:
    _close(db.execute(sql, params))


def _one(db: Any, sql: str, params: tuple = ()) -> dict | None:
    cursor = db.execute(sql, params)
    try:
        row = cursor.fetchone()
        return dict(row) if row else None
    finally:
        cursor.close()


def _all(db: Any, sql: str, params: tuple = ()) -> list[dict]:
    cursor = db.execute(sql, params)
    try:
        return [dict(row) for row in cursor.fetchall()]
    finally:
        cursor.close()


def _schema_sql(engine: str) -> list[str]:
    if engine == "sqlite":
        text = "TEXT"
        short = "TEXT"
        timestamp = "TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP"
        integer = "INTEGER"
        table_suffix = ""
        matrix_unique = "UNIQUE(version_id, job_role_id, competency_id)"
        permission_unique = "UNIQUE(role_code, scope_type, scope_code, permission)"
    else:
        text = "TEXT"
        short = "VARCHAR(255)"
        timestamp = "DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP"
        integer = "INT"
        table_suffix = " ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci"
        # UUID relationships make version_id redundant for matrix uniqueness;
        # omitting it keeps the utf8mb4 key below MySQL's 3072-byte limit.
        matrix_unique = "UNIQUE(job_role_id, competency_id)"
        permission_unique = "UNIQUE KEY uq_catalog_role_permission (role_code(80), scope_type(40), scope_code(80), permission(40))"

    return [
        f"""
        CREATE TABLE IF NOT EXISTS catalog_versions (
          id {short} PRIMARY KEY,
          code {short} NOT NULL UNIQUE,
          name {short} NOT NULL,
          revision {integer} NOT NULL DEFAULT 1,
          status {short} NOT NULL DEFAULT 'draft',
          description {text},
          effective_from {short},
          effective_to {short},
          created_by {short},
          published_by {short},
          published_at {short},
          created_at {timestamp},
          updated_at {timestamp}
        ){table_suffix}
        """,
        f"""
        CREATE TABLE IF NOT EXISTS catalog_colleges (
          id {short} PRIMARY KEY,
          version_id {short} NOT NULL,
          code {short} NOT NULL,
          name {short} NOT NULL,
          description {text},
          sort_order {integer} NOT NULL DEFAULT 0,
          enabled {integer} NOT NULL DEFAULT 1,
          metadata_json {text},
          created_at {timestamp},
          UNIQUE(version_id, code),
          FOREIGN KEY (version_id) REFERENCES catalog_versions(id) ON DELETE CASCADE
        ){table_suffix}
        """,
        f"""
        CREATE TABLE IF NOT EXISTS catalog_majors (
          id {short} PRIMARY KEY,
          version_id {short} NOT NULL,
          college_id {short} NOT NULL,
          code {short} NOT NULL,
          name {short} NOT NULL,
          degree_type {short},
          description {text},
          sort_order {integer} NOT NULL DEFAULT 0,
          enabled {integer} NOT NULL DEFAULT 1,
          metadata_json {text},
          created_at {timestamp},
          UNIQUE(version_id, code),
          FOREIGN KEY (version_id) REFERENCES catalog_versions(id) ON DELETE CASCADE,
          FOREIGN KEY (college_id) REFERENCES catalog_colleges(id) ON DELETE CASCADE
        ){table_suffix}
        """,
        f"""
        CREATE TABLE IF NOT EXISTS catalog_training_directions (
          id {short} PRIMARY KEY,
          version_id {short} NOT NULL,
          major_id {short} NOT NULL,
          code {short} NOT NULL,
          name {short} NOT NULL,
          description {text},
          sort_order {integer} NOT NULL DEFAULT 0,
          enabled {integer} NOT NULL DEFAULT 1,
          metadata_json {text},
          created_at {timestamp},
          UNIQUE(version_id, code),
          FOREIGN KEY (version_id) REFERENCES catalog_versions(id) ON DELETE CASCADE,
          FOREIGN KEY (major_id) REFERENCES catalog_majors(id) ON DELETE CASCADE
        ){table_suffix}
        """,
        f"""
        CREATE TABLE IF NOT EXISTS catalog_employment_directions (
          id {short} PRIMARY KEY,
          version_id {short} NOT NULL,
          training_direction_id {short} NOT NULL,
          code {short} NOT NULL,
          name {short} NOT NULL,
          description {text},
          sort_order {integer} NOT NULL DEFAULT 0,
          enabled {integer} NOT NULL DEFAULT 1,
          metadata_json {text},
          created_at {timestamp},
          UNIQUE(version_id, code),
          FOREIGN KEY (version_id) REFERENCES catalog_versions(id) ON DELETE CASCADE,
          FOREIGN KEY (training_direction_id) REFERENCES catalog_training_directions(id) ON DELETE CASCADE
        ){table_suffix}
        """,
        f"""
        CREATE TABLE IF NOT EXISTS catalog_job_roles (
          id {short} PRIMARY KEY,
          version_id {short} NOT NULL,
          employment_direction_id {short} NOT NULL,
          code {short} NOT NULL,
          name {short} NOT NULL,
          description {text},
          sort_order {integer} NOT NULL DEFAULT 0,
          enabled {integer} NOT NULL DEFAULT 1,
          metadata_json {text},
          created_at {timestamp},
          UNIQUE(version_id, code),
          FOREIGN KEY (version_id) REFERENCES catalog_versions(id) ON DELETE CASCADE,
          FOREIGN KEY (employment_direction_id) REFERENCES catalog_employment_directions(id) ON DELETE CASCADE
        ){table_suffix}
        """,
        f"""
        CREATE TABLE IF NOT EXISTS catalog_competencies (
          id {short} PRIMARY KEY,
          version_id {short} NOT NULL,
          code {short} NOT NULL,
          name {short} NOT NULL,
          category {short} NOT NULL,
          description {text},
          measurement_hint {text},
          sort_order {integer} NOT NULL DEFAULT 0,
          enabled {integer} NOT NULL DEFAULT 1,
          metadata_json {text},
          created_at {timestamp},
          UNIQUE(version_id, code),
          FOREIGN KEY (version_id) REFERENCES catalog_versions(id) ON DELETE CASCADE
        ){table_suffix}
        """,
        f"""
        CREATE TABLE IF NOT EXISTS catalog_job_competencies (
          id {short} PRIMARY KEY,
          version_id {short} NOT NULL,
          job_role_id {short} NOT NULL,
          competency_id {short} NOT NULL,
          required_level {integer} NOT NULL DEFAULT 1,
          weight {integer} NOT NULL DEFAULT 1,
          required {integer} NOT NULL DEFAULT 1,
          evidence_hint {text},
          created_at {timestamp},
          {matrix_unique},
          FOREIGN KEY (version_id) REFERENCES catalog_versions(id) ON DELETE CASCADE,
          FOREIGN KEY (job_role_id) REFERENCES catalog_job_roles(id) ON DELETE CASCADE,
          FOREIGN KEY (competency_id) REFERENCES catalog_competencies(id) ON DELETE CASCADE
        ){table_suffix}
        """,
        f"""
        CREATE TABLE IF NOT EXISTS catalog_role_permissions (
          id {short} PRIMARY KEY,
          role_code {short} NOT NULL,
          scope_type {short} NOT NULL DEFAULT 'catalog',
          scope_code {short} NOT NULL DEFAULT '*',
          permission {short} NOT NULL,
          created_at {timestamp},
          {permission_unique}
        ){table_suffix}
        """,
        f"""
        CREATE TABLE IF NOT EXISTS catalog_job_aliases (
          id {short} PRIMARY KEY,
          version_id {short} NOT NULL,
          job_role_id {short} NOT NULL,
          alias_name {short} NOT NULL,
          normalized_name {short} NOT NULL,
          source {short} NOT NULL DEFAULT 'manual',
          enabled {integer} NOT NULL DEFAULT 1,
          created_by {short},
          created_at {timestamp},
          UNIQUE(version_id, normalized_name),
          FOREIGN KEY (version_id) REFERENCES catalog_versions(id) ON DELETE CASCADE,
          FOREIGN KEY (job_role_id) REFERENCES catalog_job_roles(id) ON DELETE CASCADE
        ){table_suffix}
        """,
        f"""
        CREATE TABLE IF NOT EXISTS catalog_job_suggestions (
          id {short} PRIMARY KEY,
          suggested_name {short} NOT NULL,
          normalized_name {short} NOT NULL UNIQUE,
          nearest_job_role_id {short},
          match_confidence {integer} NOT NULL DEFAULT 0,
          occurrence_count {integer} NOT NULL DEFAULT 1,
          source_type {short} NOT NULL DEFAULT 'resume_ai',
          first_provider {short},
          last_provider {short},
          review_status {short} NOT NULL DEFAULT 'pending',
          review_note {text},
          reviewed_by {short},
          reviewed_at {short},
          merged_job_role_id {short},
          created_at {timestamp},
          updated_at {timestamp}
        ){table_suffix}
        """,
        f"""
        CREATE TABLE IF NOT EXISTS catalog_job_suggestion_drafts (
          id {short} PRIMARY KEY,
          suggestion_id {short} NOT NULL UNIQUE,
          version_id {short} NOT NULL,
          job_role_id {short} NOT NULL,
          created_by {short},
          created_at {timestamp},
          FOREIGN KEY (suggestion_id) REFERENCES catalog_job_suggestions(id) ON DELETE CASCADE,
          FOREIGN KEY (version_id) REFERENCES catalog_versions(id) ON DELETE CASCADE,
          FOREIGN KEY (job_role_id) REFERENCES catalog_job_roles(id) ON DELETE CASCADE
        ){table_suffix}
        """,
        f"""
        CREATE TABLE IF NOT EXISTS catalog_import_jobs (
          id {short} PRIMARY KEY,
          version_id {short} NOT NULL,
          filename {short} NOT NULL,
          file_hash {short} NOT NULL,
          import_mode {short} NOT NULL DEFAULT 'merge',
          status {short} NOT NULL DEFAULT 'pending',
          total_rows {integer} NOT NULL DEFAULT 0,
          imported_rows {integer} NOT NULL DEFAULT 0,
          error_rows {integer} NOT NULL DEFAULT 0,
          error_json {text},
          created_by {short},
          created_at {timestamp},
          finished_at {short},
          FOREIGN KEY (version_id) REFERENCES catalog_versions(id) ON DELETE CASCADE
        ){table_suffix}
        """,
    ]


def ensure_catalog_schema(db: Any, engine: str) -> None:
    for statement in _schema_sql(engine):
        _execute(db, statement)
    indexes = [
        ("idx_catalog_versions_status", "catalog_versions", "status"),
        ("idx_catalog_majors_college", "catalog_majors", "college_id"),
        ("idx_catalog_training_major", "catalog_training_directions", "major_id"),
        ("idx_catalog_employment_training", "catalog_employment_directions", "training_direction_id"),
        ("idx_catalog_jobs_employment", "catalog_job_roles", "employment_direction_id"),
        ("idx_catalog_matrix_job", "catalog_job_competencies", "job_role_id"),
        ("idx_catalog_alias_job", "catalog_job_aliases", "job_role_id"),
        ("idx_catalog_suggestion_status", "catalog_job_suggestions", "review_status"),
        ("idx_catalog_suggestion_draft_job", "catalog_job_suggestion_drafts", "job_role_id"),
        ("idx_catalog_import_version", "catalog_import_jobs", "version_id"),
    ]
    for name, table, columns in indexes:
        try:
            _execute(db, f"CREATE INDEX {name} ON {table} ({columns})")
        except Exception as exc:
            if "exist" not in str(exc).lower() and "duplicate" not in str(exc).lower():
                raise
    db.commit()


PILOT_DATA = {
    "colleges": [("CS", "计算机学院", "计算机类专业试点学院")],
    "majors": [
        ("080901", "CS", "计算机科学与技术", "工学"),
        ("080902", "CS", "软件工程", "工学"),
        ("080910T", "CS", "数据科学与大数据技术", "工学"),
    ],
    "training_directions": [
        ("CS-AI", "080901", "人工智能与智能系统"),
        ("CS-SYSTEM", "080901", "计算机系统与网络"),
        ("SE-ENGINEERING", "080902", "软件开发与工程管理"),
        ("DS-DATA", "080910T", "数据工程与智能分析"),
    ],
    "employment_directions": [
        ("AI-RD", "CS-AI", "人工智能研发"),
        ("INFRA", "CS-SYSTEM", "基础设施与云计算"),
        ("APP-DEV", "SE-ENGINEERING", "应用软件开发"),
        ("DATA-PLATFORM", "DS-DATA", "数据平台与分析"),
    ],
    "job_roles": [
        ("ML-ENGINEER", "AI-RD", "机器学习工程师"),
        ("DEVOPS-ENGINEER", "INFRA", "DevOps 工程师"),
        ("BACKEND-ENGINEER", "APP-DEV", "后端开发工程师"),
        ("FRONTEND-ENGINEER", "APP-DEV", "前端开发工程师"),
        ("DATA-ENGINEER", "DATA-PLATFORM", "数据工程师"),
    ],
    "competencies": [
        ("PROGRAMMING", "编程与代码质量", "专业基础"),
        ("ALGORITHM", "数据结构与算法", "专业基础"),
        ("DATABASE", "数据库设计与应用", "工程能力"),
        ("SYSTEM-DESIGN", "系统设计", "工程能力"),
        ("TESTING", "测试与质量保障", "工程能力"),
        ("DEVOPS", "持续交付与运维", "工程能力"),
        ("DATA-PIPELINE", "数据管道建设", "专业能力"),
        ("ML-MODELING", "机器学习建模", "专业能力"),
        ("COMMUNICATION", "沟通与协作", "通用能力"),
        ("PROBLEM-SOLVING", "问题分析与解决", "通用能力"),
    ],
    "matrix": {
        "ML-ENGINEER": [("PROGRAMMING", 4, 4), ("ALGORITHM", 4, 4), ("ML-MODELING", 5, 5), ("DATA-PIPELINE", 3, 3), ("COMMUNICATION", 3, 2)],
        "DEVOPS-ENGINEER": [("PROGRAMMING", 3, 3), ("SYSTEM-DESIGN", 4, 4), ("DEVOPS", 5, 5), ("PROBLEM-SOLVING", 4, 4), ("COMMUNICATION", 3, 3)],
        "BACKEND-ENGINEER": [("PROGRAMMING", 5, 5), ("ALGORITHM", 3, 3), ("DATABASE", 4, 4), ("SYSTEM-DESIGN", 4, 4), ("TESTING", 3, 3), ("COMMUNICATION", 3, 2)],
        "FRONTEND-ENGINEER": [("PROGRAMMING", 5, 5), ("SYSTEM-DESIGN", 3, 3), ("TESTING", 3, 3), ("COMMUNICATION", 4, 3), ("PROBLEM-SOLVING", 4, 4)],
        "DATA-ENGINEER": [("PROGRAMMING", 4, 4), ("DATABASE", 5, 5), ("SYSTEM-DESIGN", 4, 4), ("DATA-PIPELINE", 5, 5), ("PROBLEM-SOLVING", 4, 3)],
    },
}


def _seed_default_permissions(db: Any) -> None:
    for role, permissions in {
        "super_admin": ("read", "write", "import", "publish"),
        "operations": ("read", "write", "import"),
        "reviewer": ("read",),
    }.items():
        for permission in permissions:
            existing = _one(db, """SELECT id FROM catalog_role_permissions
                WHERE role_code = ? AND scope_type = 'catalog' AND scope_code = '*' AND permission = ?""", (role, permission))
            if not existing:
                _execute(db, "INSERT INTO catalog_role_permissions (id, role_code, scope_type, scope_code, permission) VALUES (?, ?, 'catalog', '*', ?)", (_id(), role, permission))


def seed_computer_pilot(db: Any) -> str:
    existing = _one(db, "SELECT id FROM catalog_versions WHERE code = ?", ("computer-pilot-v1",))
    if existing:
        _seed_default_permissions(db)
        _seed_pilot_aliases(db, existing["id"])
        _refresh_pending_suggestion_hints(db)
        db.commit()
        return existing["id"]
    version_id = _id()
    _execute(
        db,
        """INSERT INTO catalog_versions
           (id, code, name, revision, status, description, effective_from, published_at)
           VALUES (?, ?, ?, 1, 'published', ?, ?, ?)""",
        (version_id, "computer-pilot-v1", "计算机类职业能力目录 V1", "全校通用架构的计算机类专业试点数据", "2026-07-01", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")),
    )
    ids: dict[str, str] = {}
    for order, (code, name, description) in enumerate(PILOT_DATA["colleges"]):
        ids[code] = _id()
        _execute(db, "INSERT INTO catalog_colleges (id, version_id, code, name, description, sort_order, metadata_json) VALUES (?, ?, ?, ?, ?, ?, ?)", (ids[code], version_id, code, name, description, order, "{}"))
    for order, (code, parent, name, degree) in enumerate(PILOT_DATA["majors"]):
        ids[code] = _id()
        _execute(db, "INSERT INTO catalog_majors (id, version_id, college_id, code, name, degree_type, sort_order, metadata_json) VALUES (?, ?, ?, ?, ?, ?, ?, ?)", (ids[code], version_id, ids[parent], code, name, degree, order, "{}"))
    for table_key, table_name, parent_column in (
        ("training_directions", "catalog_training_directions", "major_id"),
        ("employment_directions", "catalog_employment_directions", "training_direction_id"),
        ("job_roles", "catalog_job_roles", "employment_direction_id"),
    ):
        for order, (code, parent, name) in enumerate(PILOT_DATA[table_key]):
            ids[code] = _id()
            _execute(db, f"INSERT INTO {table_name} (id, version_id, {parent_column}, code, name, sort_order, metadata_json) VALUES (?, ?, ?, ?, ?, ?, ?)", (ids[code], version_id, ids[parent], code, name, order, "{}"))
    for order, (code, name, category) in enumerate(PILOT_DATA["competencies"]):
        ids[code] = _id()
        _execute(db, "INSERT INTO catalog_competencies (id, version_id, code, name, category, sort_order, metadata_json) VALUES (?, ?, ?, ?, ?, ?, ?)", (ids[code], version_id, code, name, category, order, "{}"))
    for job_code, rows in PILOT_DATA["matrix"].items():
        for competency_code, level, weight in rows:
            _execute(db, "INSERT INTO catalog_job_competencies (id, version_id, job_role_id, competency_id, required_level, weight, required) VALUES (?, ?, ?, ?, ?, ?, 1)", (_id(), version_id, ids[job_code], ids[competency_code], level, weight))
    _seed_default_permissions(db)
    _seed_pilot_aliases(db, version_id)
    _refresh_pending_suggestion_hints(db)
    db.commit()
    return version_id


def normalize_job_name(value: str) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).strip().lower()
    text = re.sub(r"(?:高级|中级|初级|资深|实习|校招|社招|专家|负责人|应届|junior|senior|intern|lead)", "", text, flags=re.I)
    text = re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", text)
    return text


def _seed_pilot_aliases(db: Any, version_id: str) -> None:
    aliases = {
        "BACKEND-ENGINEER": ("后端工程师", "后端开发", "Java 后端工程师", "服务端开发工程师"),
        "FRONTEND-ENGINEER": ("前端工程师", "Web 前端工程师", "Web 开发工程师"),
        "ML-ENGINEER": ("机器学习开发工程师", "Machine Learning Engineer", "ML Engineer"),
        "DEVOPS-ENGINEER": ("DevOps 开发工程师",),
        "DATA-ENGINEER": ("大数据工程师", "数据开发工程师"),
    }
    for job_code, names in aliases.items():
        job = _one(db, "SELECT id FROM catalog_job_roles WHERE version_id = ? AND code = ?", (version_id, job_code))
        if not job:
            continue
        for alias_name in names:
            normalized = normalize_job_name(alias_name)
            if not _one(db, "SELECT id FROM catalog_job_aliases WHERE version_id = ? AND normalized_name = ?", (version_id, normalized)):
                _execute(db, """INSERT INTO catalog_job_aliases
                    (id, version_id, job_role_id, alias_name, normalized_name, source)
                    VALUES (?, ?, ?, ?, ?, 'seed')""", (_id(), version_id, job["id"], alias_name, normalized))


JOB_DOMAIN_KEYWORDS = {
    "ML-ENGINEER": ("人工智能", "ai", "机器学习", "深度学习", "大模型", "算法", "模型", "nlp", "视觉"),
    "DEVOPS-ENGINEER": ("devops", "运维", "云计算", "云平台", "基础设施", "网络", "安全", "渗透", "攻防", "sre", "容器", "kubernetes"),
    "BACKEND-ENGINEER": ("后端", "服务端", "java", "golang", "微服务", "接口"),
    "FRONTEND-ENGINEER": ("前端", "web", "网页", "h5", "react", "vue", "typescript"),
    "DATA-ENGINEER": ("大数据", "数据仓库", "数仓", "数据开发", "数据平台", "etl", "bi", "数据分析"),
}


def _job_domain_affinity(normalized_name: str, job_code: str) -> float:
    """Give catalog hints a domain signal so shared suffixes like “工程师” do not dominate."""
    hits = sum(1 for keyword in JOB_DOMAIN_KEYWORDS.get(job_code, ()) if normalize_job_name(keyword) in normalized_name)
    return min(0.86, 0.68 + (hits - 1) * 0.06) if hits else 0.0


def list_versions(db: Any, include_drafts: bool = False) -> list[dict]:
    where = "" if include_drafts else "WHERE status = 'published'"
    return _all(db, f"SELECT * FROM catalog_versions {where} ORDER BY revision DESC, created_at DESC")


def resolve_version(db: Any, version_code: str | None = None, include_drafts: bool = False) -> dict | None:
    if version_code:
        row = _one(db, "SELECT * FROM catalog_versions WHERE code = ?", (version_code,))
        return row if row and (include_drafts or row["status"] == "published") else None
    where = "" if include_drafts else "WHERE status = 'published'"
    return _one(db, f"SELECT * FROM catalog_versions {where} ORDER BY revision DESC, created_at DESC LIMIT 1")


def catalog_tree(db: Any, version_code: str | None = None, include_disabled: bool = False, include_drafts: bool = False) -> dict | None:
    version = resolve_version(db, version_code, include_drafts)
    if not version:
        return None
    enabled = "" if include_disabled else " AND enabled = 1"
    version_id = version["id"]
    colleges = _all(db, f"SELECT * FROM catalog_colleges WHERE version_id = ?{enabled} ORDER BY sort_order, name", (version_id,))
    majors = _all(db, f"SELECT * FROM catalog_majors WHERE version_id = ?{enabled} ORDER BY sort_order, name", (version_id,))
    training = _all(db, f"SELECT * FROM catalog_training_directions WHERE version_id = ?{enabled} ORDER BY sort_order, name", (version_id,))
    employment = _all(db, f"SELECT * FROM catalog_employment_directions WHERE version_id = ?{enabled} ORDER BY sort_order, name", (version_id,))
    jobs = _all(db, f"SELECT * FROM catalog_job_roles WHERE version_id = ?{enabled} ORDER BY sort_order, name", (version_id,))
    competencies = _all(db, f"SELECT * FROM catalog_competencies WHERE version_id = ?{enabled} ORDER BY sort_order, name", (version_id,))
    matrices = _all(db, """SELECT matrix.job_role_id, matrix.required_level, matrix.weight, matrix.required, matrix.evidence_hint,
        competencies.id, competencies.code, competencies.name, competencies.category, competencies.description
        FROM catalog_job_competencies AS matrix JOIN catalog_competencies AS competencies ON competencies.id = matrix.competency_id
        WHERE matrix.version_id = ? ORDER BY competencies.sort_order, competencies.name""", (version_id,))
    by_job: dict[str, list[dict]] = defaultdict(list)
    for item in matrices:
        by_job[item.pop("job_role_id")].append(item)
    by_employment: dict[str, list[dict]] = defaultdict(list)
    for item in jobs:
        item["competencies"] = by_job[item["id"]]
        by_employment[item.pop("employment_direction_id")].append(item)
    by_training: dict[str, list[dict]] = defaultdict(list)
    for item in employment:
        item["jobs"] = by_employment[item["id"]]
        by_training[item.pop("training_direction_id")].append(item)
    by_major: dict[str, list[dict]] = defaultdict(list)
    for item in training:
        item["employment_directions"] = by_training[item["id"]]
        by_major[item.pop("major_id")].append(item)
    by_college: dict[str, list[dict]] = defaultdict(list)
    for item in majors:
        item["training_directions"] = by_major[item["id"]]
        by_college[item.pop("college_id")].append(item)
    for item in colleges:
        item["majors"] = by_college[item["id"]]
    return {"version": version, "colleges": colleges, "competencies": competencies}


def _job_ability_matrix(db: Any, version_id: str, job_role_id: str) -> list[dict]:
    return _all(db, """SELECT competencies.id, competencies.code, competencies.name, competencies.category,
        matrix.required_level, matrix.weight, matrix.required, matrix.evidence_hint
        FROM catalog_job_competencies AS matrix
        JOIN catalog_competencies AS competencies ON competencies.id = matrix.competency_id
        WHERE matrix.version_id = ? AND matrix.job_role_id = ?
        ORDER BY matrix.weight DESC, competencies.sort_order, competencies.name""", (version_id, job_role_id))


def match_catalog_job(db: Any, suggested_name: str) -> dict:
    version = resolve_version(db)
    normalized = normalize_job_name(suggested_name)
    if not version or not normalized:
        return {"catalog_status": "external", "match_method": "none", "match_confidence": 0}
    jobs = _all(db, "SELECT id, code, name FROM catalog_job_roles WHERE version_id = ? AND enabled = 1", (version["id"],))
    aliases = _all(db, """SELECT aliases.normalized_name, jobs.id, jobs.code, jobs.name
        FROM catalog_job_aliases AS aliases
        JOIN catalog_job_roles AS jobs ON jobs.id = aliases.job_role_id
        WHERE aliases.version_id = ? AND aliases.enabled = 1 AND jobs.enabled = 1""", (version["id"],))
    for job in jobs:
        if normalize_job_name(job["name"]) == normalized:
            return {
                "catalog_status": "matched", "catalog_job_id": job["id"], "catalog_job_code": job["code"],
                "catalog_job_name": job["name"], "catalog_version": version["code"], "match_method": "exact",
                "match_confidence": 100, "ability_matrix": _job_ability_matrix(db, version["id"], job["id"]),
            }
    for alias in aliases:
        if alias["normalized_name"] == normalized:
            return {
                "catalog_status": "matched", "catalog_job_id": alias["id"], "catalog_job_code": alias["code"],
                "catalog_job_name": alias["name"], "catalog_version": version["code"], "match_method": "alias",
                "match_confidence": 100, "ability_matrix": _job_ability_matrix(db, version["id"], alias["id"]),
            }
    nearest = None
    nearest_score = 0.0
    nearest_method = "similarity_hint"
    for job in jobs:
        similarity_score = SequenceMatcher(None, normalized, normalize_job_name(job["name"])).ratio()
        domain_score = _job_domain_affinity(normalized, job["code"])
        score = max(similarity_score, domain_score)
        if score > nearest_score:
            nearest, nearest_score = job, score
            nearest_method = "semantic_hint" if domain_score > similarity_score else "similarity_hint"
    return {
        "catalog_status": "external",
        "nearest_catalog_job_id": nearest["id"] if nearest else None,
        "nearest_catalog_job_name": nearest["name"] if nearest else None,
        "catalog_version": version["code"],
        "match_method": nearest_method if nearest else "none",
        "match_confidence": round(nearest_score * 100),
    }


CATALOG_MATCH_FIELDS = {
    "catalog_status",
    "catalog_job_id",
    "catalog_job_code",
    "catalog_job_name",
    "catalog_version",
    "match_method",
    "match_confidence",
    "ability_matrix",
    "nearest_catalog_job_id",
    "nearest_catalog_job_name",
}


def refresh_resume_direction_matches(db: Any, directions: list[dict]) -> list[dict]:
    """Refresh catalog-derived fields without invoking an AI model or duplicating review-pool records."""
    refreshed: list[dict] = []
    for direction in directions:
        if not isinstance(direction, dict):
            continue
        cleaned = {key: value for key, value in direction.items() if key not in CATALOG_MATCH_FIELDS}
        refreshed.append({**cleaned, **match_catalog_job(db, str(direction.get("name") or ""))})
    return refreshed


def _refresh_pending_suggestion_hints(db: Any) -> None:
    suggestions = _all(
        db,
        "SELECT id, suggested_name FROM catalog_job_suggestions WHERE review_status = 'pending'",
    )
    for suggestion in suggestions:
        match = match_catalog_job(db, suggestion["suggested_name"])
        _execute(
            db,
            """UPDATE catalog_job_suggestions
               SET nearest_job_role_id = ?, match_confidence = ?, updated_at = CURRENT_TIMESTAMP
               WHERE id = ?""",
            (
                match.get("nearest_catalog_job_id"),
                int(match.get("match_confidence") or 0),
                suggestion["id"],
            ),
        )


def record_job_suggestion(db: Any, suggested_name: str, match: dict, provider: str | None = None) -> dict:
    normalized = normalize_job_name(suggested_name)
    if not normalized:
        raise ValueError("岗位建议名称不能为空。")
    existing = _one(db, "SELECT * FROM catalog_job_suggestions WHERE normalized_name = ?", (normalized,))
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    if existing:
        _execute(db, """UPDATE catalog_job_suggestions
            SET suggested_name = ?, nearest_job_role_id = ?, match_confidence = ?,
                occurrence_count = occurrence_count + 1, last_provider = ?, updated_at = ?
            WHERE id = ?""", (
                suggested_name[:160], match.get("nearest_catalog_job_id"), int(match.get("match_confidence") or 0),
                provider, now, existing["id"],
            ))
        suggestion_id = existing["id"]
    else:
        suggestion_id = _id()
        _execute(db, """INSERT INTO catalog_job_suggestions
            (id, suggested_name, normalized_name, nearest_job_role_id, match_confidence,
             occurrence_count, source_type, first_provider, last_provider, review_status)
            VALUES (?, ?, ?, ?, ?, 1, 'resume_ai', ?, ?, 'pending')""", (
                suggestion_id, suggested_name[:160], normalized, match.get("nearest_catalog_job_id"),
                int(match.get("match_confidence") or 0), provider, provider,
            ))
    db.commit()
    return _one(db, "SELECT * FROM catalog_job_suggestions WHERE id = ?", (suggestion_id,)) or {}


def enrich_resume_directions(db: Any, directions: list[dict], provider: str | None = None) -> list[dict]:
    enriched = []
    for direction in directions:
        item = dict(direction)
        match = match_catalog_job(db, str(item.get("name") or ""))
        item.update(match)
        if match["catalog_status"] == "external":
            suggestion = record_job_suggestion(db, str(item.get("name") or ""), match, provider)
            item["suggestion_id"] = suggestion.get("id")
            item["suggestion_status"] = suggestion.get("review_status") or "pending"
        enriched.append(item)
    return enriched


def list_job_suggestions(db: Any, status: str | None = None) -> list[dict]:
    params: tuple = ()
    where = ""
    if status:
        where = "WHERE suggestions.review_status = ?"
        params = (status,)
    return _all(db, f"""SELECT suggestions.*, nearest.name AS nearest_job_name,
        merged.name AS merged_job_name, draft_job.id AS draft_job_role_id,
        draft_job.name AS draft_job_name, draft_version.id AS draft_version_id,
        draft_version.code AS draft_version_code
        FROM catalog_job_suggestions AS suggestions
        LEFT JOIN catalog_job_roles AS nearest ON nearest.id = suggestions.nearest_job_role_id
        LEFT JOIN catalog_job_roles AS merged ON merged.id = suggestions.merged_job_role_id
        LEFT JOIN catalog_job_suggestion_drafts AS draft_link ON draft_link.suggestion_id = suggestions.id
        LEFT JOIN catalog_job_roles AS draft_job ON draft_job.id = draft_link.job_role_id
        LEFT JOIN catalog_versions AS draft_version ON draft_version.id = draft_link.version_id
        {where}
        ORDER BY CASE suggestions.review_status WHEN 'pending' THEN 0 WHEN 'approved' THEN 1 ELSE 2 END,
                 suggestions.occurrence_count DESC, suggestions.updated_at DESC""", params)


def review_job_suggestion(db: Any, suggestion_id: str, status: str, reviewed_by: str | None, note: str | None = None) -> dict | None:
    if status not in {"pending", "approved", "rejected"}:
        raise ValueError("岗位建议审核状态不正确。")
    if not _one(db, "SELECT id FROM catalog_job_suggestions WHERE id = ?", (suggestion_id,)):
        return None
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    _execute(db, """UPDATE catalog_job_suggestions
        SET review_status = ?, review_note = ?, reviewed_by = ?, reviewed_at = ?, updated_at = ?
        WHERE id = ?""", (status, str(note or "").strip() or None, reviewed_by, now, now, suggestion_id))
    db.commit()
    return _one(db, "SELECT * FROM catalog_job_suggestions WHERE id = ?", (suggestion_id,))


def merge_job_suggestion(db: Any, suggestion_id: str, job_role_id: str, reviewed_by: str | None) -> dict | None:
    suggestion = _one(db, "SELECT * FROM catalog_job_suggestions WHERE id = ?", (suggestion_id,))
    job = _one(db, """SELECT jobs.id, jobs.version_id, jobs.name FROM catalog_job_roles AS jobs
        JOIN catalog_versions AS versions ON versions.id = jobs.version_id
        WHERE jobs.id = ? AND jobs.enabled = 1 AND versions.status = 'published'""", (job_role_id,))
    if not suggestion:
        return None
    if not job:
        raise ValueError("只能合并到当前已发布目录中的有效岗位。")
    normalized = normalize_job_name(suggestion["suggested_name"])
    existing_alias = _one(db, "SELECT id FROM catalog_job_aliases WHERE version_id = ? AND normalized_name = ?", (job["version_id"], normalized))
    if not existing_alias:
        _execute(db, """INSERT INTO catalog_job_aliases
            (id, version_id, job_role_id, alias_name, normalized_name, source, created_by)
            VALUES (?, ?, ?, ?, ?, 'suggestion_merge', ?)""", (
                _id(), job["version_id"], job_role_id, suggestion["suggested_name"], normalized, reviewed_by,
            ))
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    _execute(db, """UPDATE catalog_job_suggestions
        SET review_status = 'merged', merged_job_role_id = ?, reviewed_by = ?, reviewed_at = ?, updated_at = ?
        WHERE id = ?""", (job_role_id, reviewed_by, now, now, suggestion_id))
    db.commit()
    return _one(db, "SELECT * FROM catalog_job_suggestions WHERE id = ?", (suggestion_id,))


def has_permission(db: Any, role: str, permission: str, scope_code: str = "*") -> bool:
    return bool(_one(db, """SELECT id FROM catalog_role_permissions
        WHERE role_code = ? AND permission = ? AND scope_type = 'catalog' AND scope_code IN ('*', ?)
        LIMIT 1""", (role, permission, scope_code)))


def create_version(db: Any, payload: dict, created_by: str | None) -> dict:
    code = str(payload.get("code") or "").strip()
    name = str(payload.get("name") or "").strip()
    if not code or not name:
        raise ValueError("版本编码和名称不能为空。")
    if _one(db, "SELECT id FROM catalog_versions WHERE code = ?", (code,)):
        raise ValueError("版本编码已存在。")
    revision_row = _one(db, "SELECT MAX(revision) AS revision FROM catalog_versions") or {}
    version_id = _id()
    _execute(db, """INSERT INTO catalog_versions
        (id, code, name, revision, status, description, effective_from, effective_to, created_by)
        VALUES (?, ?, ?, ?, 'draft', ?, ?, ?, ?)""", (
            version_id, code[:100], name[:160], int(revision_row.get("revision") or 0) + 1,
            str(payload.get("description") or "").strip() or None,
            str(payload.get("effective_from") or "").strip() or None,
            str(payload.get("effective_to") or "").strip() or None,
            created_by,
        ))
    db.commit()
    return _one(db, "SELECT * FROM catalog_versions WHERE id = ?", (version_id,)) or {}


def _draft_version(db: Any, version_id: str) -> dict:
    version = _one(db, "SELECT * FROM catalog_versions WHERE id = ?", (version_id,))
    if not version:
        raise ValueError("目录版本不存在。")
    if version["status"] != "draft":
        raise ValueError("只能直接编辑草稿版本。")
    return version


def _entity_spec(entity_type: str) -> dict:
    spec = CATALOG_ENTITY_SPECS.get(str(entity_type or "").strip())
    if not spec:
        raise ValueError("不支持的目录数据类型。")
    return spec


def _catalog_sort_order(value: Any) -> int:
    try:
        result = int(value or 0)
    except (TypeError, ValueError) as exc:
        raise ValueError("排序值必须是整数。") from exc
    if not -100000 <= result <= 100000:
        raise ValueError("排序值超出允许范围。")
    return result


def _catalog_enabled(value: Any) -> int:
    if isinstance(value, str):
        return 1 if value.strip().lower() in {"1", "true", "yes", "on"} else 0
    return 1 if bool(value) else 0


def _catalog_code(value: Any) -> str:
    code = str(value or "").strip()
    if not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9._-]{0,99}", code):
        raise ValueError("目录编码需以字母或数字开头，且只能包含字母、数字、点、下划线和短横线。")
    return code


def _catalog_name(value: Any) -> str:
    name = str(value or "").strip()
    if not name:
        raise ValueError("目录名称不能为空。")
    if len(name) > 160:
        raise ValueError("目录名称不能超过 160 个字符。")
    return name


def _catalog_entity_values(db: Any, version_id: str, spec: dict, payload: dict, current: dict | None = None) -> dict:
    merged = {**(current or {}), **payload}
    values: dict[str, Any] = {}
    for field in spec["fields"]:
        value = merged.get(field)
        if field == "code":
            values[field] = _catalog_code(value)
        elif field == "name":
            values[field] = _catalog_name(value)
        elif field == "sort_order":
            values[field] = _catalog_sort_order(value)
        elif field == "enabled":
            values[field] = _catalog_enabled(1 if value is None else value)
        else:
            cleaned = str(value or "").strip()
            values[field] = cleaned or None
    if spec["table"] == "catalog_competencies" and not values.get("category"):
        raise ValueError("能力类别不能为空。")
    parent_field = spec["parent_field"]
    if parent_field:
        parent_id = str(values.get(parent_field) or "").strip()
        parent = _one(
            db,
            f"SELECT id FROM {spec['parent_table']} WHERE id = ? AND version_id = ?",
            (parent_id, version_id),
        )
        if not parent:
            raise ValueError("所选上级目录不存在或不属于当前草稿。")
        values[parent_field] = parent_id
    return values


def create_catalog_entity(db: Any, version_id: str, entity_type: str, payload: dict, *, commit: bool = True) -> dict:
    _draft_version(db, version_id)
    spec = _entity_spec(entity_type)
    values = _catalog_entity_values(db, version_id, spec, payload)
    duplicate = _one(
        db,
        f"SELECT id FROM {spec['table']} WHERE version_id = ? AND code = ?",
        (version_id, values["code"]),
    )
    if duplicate:
        raise ValueError("当前目录版本中已存在相同编码。")
    item_id = _id()
    columns = ["id", "version_id", *values.keys(), "metadata_json"]
    params = (item_id, version_id, *values.values(), "{}")
    _execute(
        db,
        f"INSERT INTO {spec['table']} ({', '.join(columns)}) VALUES ({', '.join('?' for _ in columns)})",
        params,
    )
    if commit:
        db.commit()
    return _one(db, f"SELECT * FROM {spec['table']} WHERE id = ?", (item_id,)) or {}


def update_catalog_entity(db: Any, version_id: str, entity_type: str, entity_id: str, payload: dict) -> dict | None:
    _draft_version(db, version_id)
    spec = _entity_spec(entity_type)
    current = _one(
        db,
        f"SELECT * FROM {spec['table']} WHERE id = ? AND version_id = ?",
        (entity_id, version_id),
    )
    if not current:
        return None
    values = _catalog_entity_values(db, version_id, spec, payload, current)
    duplicate = _one(
        db,
        f"SELECT id FROM {spec['table']} WHERE version_id = ? AND code = ? AND id <> ?",
        (version_id, values["code"], entity_id),
    )
    if duplicate:
        raise ValueError("当前目录版本中已存在相同编码。")
    assignments = ", ".join(f"{field} = ?" for field in values)
    _execute(db, f"UPDATE {spec['table']} SET {assignments} WHERE id = ? AND version_id = ?", (*values.values(), entity_id, version_id))
    db.commit()
    return _one(db, f"SELECT * FROM {spec['table']} WHERE id = ?", (entity_id,))


def delete_catalog_entity(db: Any, version_id: str, entity_type: str, entity_id: str) -> bool:
    _draft_version(db, version_id)
    spec = _entity_spec(entity_type)
    if not _one(db, f"SELECT id FROM {spec['table']} WHERE id = ? AND version_id = ?", (entity_id, version_id)):
        return False
    _execute(db, f"DELETE FROM {spec['table']} WHERE id = ? AND version_id = ?", (entity_id, version_id))
    db.commit()
    return True


def replace_job_competencies(
    db: Any,
    version_id: str,
    job_role_id: str,
    items: list[dict],
    *,
    commit: bool = True,
) -> list[dict]:
    _draft_version(db, version_id)
    job = _one(db, "SELECT id FROM catalog_job_roles WHERE id = ? AND version_id = ?", (job_role_id, version_id))
    if not job:
        raise ValueError("岗位不存在或不属于当前草稿。")
    normalized: list[dict] = []
    seen: set[str] = set()
    for item in items:
        competency_id = str(item.get("competency_id") or "").strip()
        if not competency_id or competency_id in seen:
            raise ValueError("岗位能力配置包含空值或重复能力。")
        competency = _one(
            db,
            "SELECT id FROM catalog_competencies WHERE id = ? AND version_id = ? AND enabled = 1",
            (competency_id, version_id),
        )
        if not competency:
            raise ValueError("岗位能力引用了不存在或已停用的能力。")
        try:
            required_level = int(item.get("required_level"))
            weight = int(item.get("weight"))
        except (TypeError, ValueError) as exc:
            raise ValueError("能力等级和权重必须是 1-5 的整数。") from exc
        if not 1 <= required_level <= 5 or not 1 <= weight <= 5:
            raise ValueError("能力等级和权重必须是 1-5 的整数。")
        normalized.append(
            {
                "competency_id": competency_id,
                "required_level": required_level,
                "weight": weight,
                "required": _catalog_enabled(item.get("required", True)),
                "evidence_hint": str(item.get("evidence_hint") or "").strip() or None,
            }
        )
        seen.add(competency_id)
    _execute(db, "DELETE FROM catalog_job_competencies WHERE version_id = ? AND job_role_id = ?", (version_id, job_role_id))
    for item in normalized:
        _execute(
            db,
            """INSERT INTO catalog_job_competencies
               (id, version_id, job_role_id, competency_id, required_level, weight, required, evidence_hint)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                _id(),
                version_id,
                job_role_id,
                item["competency_id"],
                item["required_level"],
                item["weight"],
                item["required"],
                item["evidence_hint"],
            ),
        )
    if commit:
        db.commit()
    return _job_ability_matrix(db, version_id, job_role_id)


def approve_job_suggestion_to_draft(
    db: Any,
    suggestion_id: str,
    version_id: str,
    payload: dict,
    reviewed_by: str | None,
) -> dict | None:
    suggestion = _one(db, "SELECT * FROM catalog_job_suggestions WHERE id = ?", (suggestion_id,))
    if not suggestion:
        return None
    if suggestion["review_status"] == "merged":
        raise ValueError("已合并为别名的岗位建议不能再新增到草稿。")
    _draft_version(db, version_id)
    matrix = payload.get("competencies")
    if not isinstance(matrix, list) or not matrix:
        raise ValueError("批准新增岗位时必须至少配置一项能力要求。")
    job_payload = {
        "employment_direction_id": payload.get("employment_direction_id"),
        "code": payload.get("code"),
        "name": payload.get("name") or suggestion["suggested_name"],
        "description": payload.get("description"),
        "sort_order": payload.get("sort_order", 0),
        "enabled": True,
    }
    if hasattr(db, "begin"):
        db.begin()
    try:
        job = create_catalog_entity(db, version_id, "job_roles", job_payload, commit=False)
        competencies = replace_job_competencies(db, version_id, job["id"], matrix, commit=False)
        existing_link = _one(db, "SELECT id FROM catalog_job_suggestion_drafts WHERE suggestion_id = ?", (suggestion_id,))
        if existing_link:
            _execute(
                db,
                "UPDATE catalog_job_suggestion_drafts SET version_id = ?, job_role_id = ?, created_by = ? WHERE suggestion_id = ?",
                (version_id, job["id"], reviewed_by, suggestion_id),
            )
        else:
            _execute(
                db,
                """INSERT INTO catalog_job_suggestion_drafts
                   (id, suggestion_id, version_id, job_role_id, created_by) VALUES (?, ?, ?, ?, ?)""",
                (_id(), suggestion_id, version_id, job["id"], reviewed_by),
            )
        now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        _execute(
            db,
            """UPDATE catalog_job_suggestions
               SET review_status = 'approved', review_note = ?, reviewed_by = ?, reviewed_at = ?, updated_at = ?
               WHERE id = ?""",
            (str(payload.get("note") or "").strip() or "已纳入目录草稿", reviewed_by, now, now, suggestion_id),
        )
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {"suggestion": _one(db, "SELECT * FROM catalog_job_suggestions WHERE id = ?", (suggestion_id,)), "job_role": job, "competencies": competencies}


def publish_version(db: Any, version_id: str, published_by: str | None) -> dict | None:
    row = _one(db, "SELECT * FROM catalog_versions WHERE id = ?", (version_id,))
    if not row:
        return None
    required_tables = (
        "catalog_colleges",
        "catalog_majors",
        "catalog_training_directions",
        "catalog_employment_directions",
        "catalog_job_roles",
        "catalog_competencies",
        "catalog_job_competencies",
    )
    missing = [table for table in required_tables if not _one(db, f"SELECT id FROM {table} WHERE version_id = ? LIMIT 1", (version_id,))]
    if missing:
        raise ValueError("目录层级或能力矩阵不完整，暂不能发布。")
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    _execute(db, "UPDATE catalog_versions SET status = 'archived', updated_at = ? WHERE status = 'published' AND id <> ?", (now, version_id))
    _execute(db, "UPDATE catalog_versions SET status = 'published', published_by = ?, published_at = ?, updated_at = ? WHERE id = ?", (published_by, now, now, version_id))
    db.commit()
    return _one(db, "SELECT * FROM catalog_versions WHERE id = ?", (version_id,))


SHEET_SPECS = {
    "学院": ("catalog_colleges", None, ("编码", "名称")),
    "专业": ("catalog_majors", ("college_id", "学院编码"), ("编码", "名称")),
    "培养方向": ("catalog_training_directions", ("major_id", "专业编码"), ("编码", "名称")),
    "就业方向": ("catalog_employment_directions", ("training_direction_id", "培养方向编码"), ("编码", "名称")),
    "岗位": ("catalog_job_roles", ("employment_direction_id", "就业方向编码"), ("编码", "名称")),
    "能力": ("catalog_competencies", None, ("编码", "名称", "类别")),
}


def parse_excel_workbook(content: bytes) -> dict[str, list[dict]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("服务端未安装 openpyxl，无法读取 Excel。") from exc
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Excel 文件无法读取。") from exc
    result: dict[str, list[dict]] = {}
    for sheet_name, (_table, _parent, required) in SHEET_SPECS.items():
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"缺少工作表：{sheet_name}。")
        rows = list(workbook[sheet_name].iter_rows(values_only=True))
        if not rows:
            raise ValueError(f"工作表 {sheet_name} 为空。")
        headers = [str(value or "").strip() for value in rows[0]]
        for column in required + ((_parent[1],) if _parent else ()):
            if column not in headers:
                raise ValueError(f"工作表 {sheet_name} 缺少列：{column}。")
        result[sheet_name] = [
            {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
            for row in rows[1:] if any(value is not None and str(value).strip() for value in row)
        ]
    if "岗位能力" not in workbook.sheetnames:
        raise ValueError("缺少工作表：岗位能力。")
    rows = list(workbook["岗位能力"].iter_rows(values_only=True))
    headers = [str(value or "").strip() for value in (rows[0] if rows else [])]
    for column in ("岗位编码", "能力编码", "要求等级", "权重"):
        if column not in headers:
            raise ValueError(f"工作表 岗位能力 缺少列：{column}。")
    result["岗位能力"] = [{headers[i]: value for i, value in enumerate(row) if i < len(headers)} for row in rows[1:] if any(value is not None and str(value).strip() for value in row)]
    return result


def import_catalog_excel(db: Any, version_id: str, filename: str, content: bytes, mode: str, created_by: str | None) -> dict:
    if mode not in IMPORT_MODES:
        raise ValueError("导入模式仅支持 merge 或 replace。")
    version = _one(db, "SELECT * FROM catalog_versions WHERE id = ?", (version_id,))
    if not version:
        raise ValueError("目录版本不存在。")
    if version["status"] != "draft":
        raise ValueError("只能向草稿版本导入数据。")
    import_id = _id()
    total = 0
    _execute(db, """INSERT INTO catalog_import_jobs
        (id, version_id, filename, file_hash, import_mode, status, total_rows, created_by)
        VALUES (?, ?, ?, ?, ?, 'pending', 0, ?)""", (import_id, version_id, filename[:255], hashlib.sha256(content).hexdigest(), mode, created_by))
    db.commit()
    try:
        parsed = parse_excel_workbook(content)
        total = sum(len(rows) for rows in parsed.values())
        _execute(db, "UPDATE catalog_import_jobs SET status = 'running', total_rows = ? WHERE id = ?", (total, import_id))
        db.commit()
        if hasattr(db, "begin"):
            db.begin()
        if mode == "replace":
            for table in ("catalog_job_competencies", "catalog_job_roles", "catalog_employment_directions", "catalog_training_directions", "catalog_majors", "catalog_colleges", "catalog_competencies"):
                _execute(db, f"DELETE FROM {table} WHERE version_id = ?", (version_id,))
        ids: dict[str, str] = {}
        for sheet_name, (table, parent, _required) in SHEET_SPECS.items():
            for order, row in enumerate(parsed[sheet_name]):
                code = str(row.get("编码") or "").strip()
                name = str(row.get("名称") or "").strip()
                if not code or not name:
                    raise ValueError(f"{sheet_name} 第 {order + 2} 行编码或名称为空。")
                existing = _one(db, f"SELECT id FROM {table} WHERE version_id = ? AND code = ?", (version_id, code))
                item_id = existing["id"] if existing else _id()
                ids[code] = item_id
                category = str(row.get("类别") or "").strip()
                if parent:
                    parent_code = str(row.get(parent[1]) or "").strip()
                    parent_id = ids.get(parent_code)
                    if not parent_id:
                        parent_table = {"college_id": "catalog_colleges", "major_id": "catalog_majors", "training_direction_id": "catalog_training_directions", "employment_direction_id": "catalog_employment_directions"}[parent[0]]
                        found = _one(db, f"SELECT id FROM {parent_table} WHERE version_id = ? AND code = ?", (version_id, parent_code))
                        parent_id = found["id"] if found else None
                    if not parent_id:
                        raise ValueError(f"{sheet_name} 第 {order + 2} 行引用的上级编码不存在：{parent_code}。")
                    if existing:
                        _execute(db, f"UPDATE {table} SET {parent[0]} = ?, name = ?, description = ?, sort_order = ?, enabled = 1 WHERE id = ?", (parent_id, name, str(row.get("说明") or "").strip() or None, order, item_id))
                    else:
                        _execute(db, f"INSERT INTO {table} (id, version_id, {parent[0]}, code, name, description, sort_order, metadata_json) VALUES (?, ?, ?, ?, ?, ?, ?, ?)", (item_id, version_id, parent_id, code, name, str(row.get("说明") or "").strip() or None, order, "{}"))
                elif table == "catalog_competencies":
                    if not category:
                        raise ValueError(f"能力 第 {order + 2} 行类别为空。")
                    if existing:
                        _execute(db, "UPDATE catalog_competencies SET name = ?, category = ?, description = ?, sort_order = ?, enabled = 1 WHERE id = ?", (name, category, str(row.get("说明") or "").strip() or None, order, item_id))
                    else:
                        _execute(db, "INSERT INTO catalog_competencies (id, version_id, code, name, category, description, sort_order, metadata_json) VALUES (?, ?, ?, ?, ?, ?, ?, ?)", (item_id, version_id, code, name, category, str(row.get("说明") or "").strip() or None, order, "{}"))
                else:
                    if existing:
                        _execute(db, f"UPDATE {table} SET name = ?, description = ?, sort_order = ?, enabled = 1 WHERE id = ?", (name, str(row.get("说明") or "").strip() or None, order, item_id))
                    else:
                        _execute(db, f"INSERT INTO {table} (id, version_id, code, name, description, sort_order, metadata_json) VALUES (?, ?, ?, ?, ?, ?, ?)", (item_id, version_id, code, name, str(row.get("说明") or "").strip() or None, order, "{}"))
        for order, row in enumerate(parsed["岗位能力"]):
            job_code = str(row.get("岗位编码") or "").strip()
            competency_code = str(row.get("能力编码") or "").strip()
            job = _one(db, "SELECT id FROM catalog_job_roles WHERE version_id = ? AND code = ?", (version_id, job_code))
            competency = _one(db, "SELECT id FROM catalog_competencies WHERE version_id = ? AND code = ?", (version_id, competency_code))
            if not job or not competency:
                raise ValueError(f"岗位能力 第 {order + 2} 行引用的岗位或能力不存在。")
            try:
                level, weight = int(row.get("要求等级")), int(row.get("权重"))
            except (TypeError, ValueError) as exc:
                raise ValueError(f"岗位能力 第 {order + 2} 行等级或权重不是整数。") from exc
            if not 1 <= level <= 5 or not 1 <= weight <= 5:
                raise ValueError(f"岗位能力 第 {order + 2} 行等级和权重必须为 1-5。")
            existing = _one(db, "SELECT id FROM catalog_job_competencies WHERE version_id = ? AND job_role_id = ? AND competency_id = ?", (version_id, job["id"], competency["id"]))
            if existing:
                _execute(db, "UPDATE catalog_job_competencies SET required_level = ?, weight = ? WHERE id = ?", (level, weight, existing["id"]))
            else:
                _execute(db, "INSERT INTO catalog_job_competencies (id, version_id, job_role_id, competency_id, required_level, weight, required) VALUES (?, ?, ?, ?, ?, ?, 1)", (_id(), version_id, job["id"], competency["id"], level, weight))
        now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        _execute(db, "UPDATE catalog_import_jobs SET status = 'succeeded', imported_rows = ?, finished_at = ? WHERE id = ?", (total, now, import_id))
        db.commit()
    except Exception as exc:
        db.rollback()
        now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        _execute(db, "UPDATE catalog_import_jobs SET status = 'failed', error_rows = ?, error_json = ?, finished_at = ? WHERE id = ?", (total, json.dumps({"error": str(exc)}, ensure_ascii=False), now, import_id))
        db.commit()
        raise
    return _one(db, "SELECT * FROM catalog_import_jobs WHERE id = ?", (import_id,)) or {}
