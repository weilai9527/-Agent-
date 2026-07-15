from __future__ import annotations

import io
import sqlite3
import unittest

from openpyxl import Workbook

from shared.career_catalog import (
    approve_job_suggestion_to_draft,
    catalog_tree,
    create_catalog_entity,
    create_version,
    delete_catalog_entity,
    ensure_catalog_schema,
    enrich_resume_directions,
    has_permission,
    import_catalog_excel,
    list_job_suggestions,
    list_versions,
    match_catalog_job,
    merge_job_suggestion,
    publish_version,
    refresh_resume_direction_matches,
    review_job_suggestion,
    seed_computer_pilot,
    update_catalog_entity,
)


class MemoryDatabase:
    def __init__(self):
        self.connection = sqlite3.connect(":memory:")
        self.connection.row_factory = sqlite3.Row
        self.connection.execute("PRAGMA foreign_keys = ON")

    def execute(self, sql: str, params: tuple = ()):
        return self.connection.execute(sql, params)

    def commit(self):
        self.connection.commit()

    def rollback(self):
        self.connection.rollback()


def build_workbook() -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = {
        "学院": (["编码", "名称", "说明"], [["BUS", "商学院", "导入测试"]]),
        "专业": (["编码", "名称", "学院编码"], [["1202", "工商管理", "BUS"]]),
        "培养方向": (["编码", "名称", "专业编码"], [["BUS-DIGITAL", "数字化管理", "1202"]]),
        "就业方向": (["编码", "名称", "培养方向编码"], [["PRODUCT", "数字产品", "BUS-DIGITAL"]]),
        "岗位": (["编码", "名称", "就业方向编码"], [["PRODUCT-MANAGER", "产品经理", "PRODUCT"]]),
        "能力": (["编码", "名称", "类别"], [["USER-RESEARCH", "用户研究", "专业能力"]]),
        "岗位能力": (["岗位编码", "能力编码", "要求等级", "权重"], [["PRODUCT-MANAGER", "USER-RESEARCH", 4, 5]]),
    }
    for name, (headers, rows) in sheets.items():
        sheet = workbook.create_sheet(name)
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


class CareerCatalogTests(unittest.TestCase):
    def setUp(self):
        self.db = MemoryDatabase()
        ensure_catalog_schema(self.db, "sqlite")
        self.pilot_id = seed_computer_pilot(self.db)

    def test_pilot_tree_is_data_driven_and_contains_ability_matrix(self):
        tree = catalog_tree(self.db)
        self.assertEqual(tree["version"]["code"], "computer-pilot-v1")
        self.assertEqual(len(tree["colleges"]), 1)
        self.assertEqual(len(tree["colleges"][0]["majors"]), 3)
        jobs = [
            job
            for major in tree["colleges"][0]["majors"]
            for training in major["training_directions"]
            for employment in training["employment_directions"]
            for job in employment["jobs"]
        ]
        self.assertEqual(len(jobs), 5)
        self.assertTrue(all(job["competencies"] for job in jobs))

    def test_permissions_are_stored_as_data(self):
        self.assertTrue(has_permission(self.db, "operations", "import"))
        self.assertFalse(has_permission(self.db, "reviewer", "import"))
        self.assertTrue(has_permission(self.db, "reviewer", "read"))

    def test_excel_import_can_add_a_new_major_domain_without_code_changes(self):
        version = create_version(self.db, {"code": "school-v2", "name": "全校目录 V2"}, "admin-1")
        job = import_catalog_excel(self.db, version["id"], "catalog.xlsx", build_workbook(), "replace", "admin-1")
        self.assertEqual(job["status"], "succeeded")
        self.assertEqual(job["imported_rows"], 7)

        draft_tree = catalog_tree(self.db, "school-v2", include_drafts=True)
        self.assertEqual(draft_tree["colleges"][0]["name"], "商学院")
        matrix = draft_tree["colleges"][0]["majors"][0]["training_directions"][0]["employment_directions"][0]["jobs"][0]["competencies"]
        self.assertEqual(matrix[0]["code"], "USER-RESEARCH")
        self.assertEqual(matrix[0]["required_level"], 4)

        publish_version(self.db, version["id"], "admin-1")
        published = list_versions(self.db)
        self.assertEqual([item["code"] for item in published], ["school-v2"])

    def test_empty_draft_cannot_be_published(self):
        version = create_version(self.db, {"code": "empty-v2", "name": "空目录"}, "admin-1")
        with self.assertRaisesRegex(ValueError, "不完整"):
            publish_version(self.db, version["id"], "admin-1")

    def test_invalid_excel_is_recorded_as_a_failed_import(self):
        version = create_version(self.db, {"code": "invalid-v2", "name": "错误导入"}, "admin-1")
        workbook = Workbook()
        output = io.BytesIO()
        workbook.save(output)
        with self.assertRaisesRegex(ValueError, "缺少工作表"):
            import_catalog_excel(self.db, version["id"], "broken.xlsx", output.getvalue(), "merge", "admin-1")
        row = self.db.execute("SELECT status, error_json FROM catalog_import_jobs WHERE version_id = ?", (version["id"],)).fetchone()
        self.assertEqual(row["status"], "failed")
        self.assertIn("缺少工作表", row["error_json"])

    def test_formal_and_alias_job_names_link_to_the_ability_matrix(self):
        exact = match_catalog_job(self.db, "后端开发工程师")
        alias = match_catalog_job(self.db, "Java 后端工程师")

        self.assertEqual(exact["catalog_status"], "matched")
        self.assertEqual(exact["match_method"], "exact")
        self.assertTrue(exact["ability_matrix"])
        self.assertEqual(alias["catalog_status"], "matched")
        self.assertEqual(alias["match_method"], "alias")
        self.assertEqual(alias["catalog_job_id"], exact["catalog_job_id"])

    def test_external_job_is_aggregated_in_the_review_pool_without_resume_data(self):
        directions = [{"name": "大模型应用开发工程师", "score": 88, "reason": "有 AI 项目经验"}]
        first = enrich_resume_directions(self.db, directions, "qwen")
        second = enrich_resume_directions(self.db, directions, "openai")
        suggestions = list_job_suggestions(self.db, "pending")

        self.assertEqual(first[0]["catalog_status"], "external")
        self.assertEqual(first[0]["suggestion_status"], "pending")
        self.assertEqual(second[0]["suggestion_id"], first[0]["suggestion_id"])
        self.assertEqual(len(suggestions), 1)
        self.assertEqual(suggestions[0]["occurrence_count"], 2)
        self.assertEqual(suggestions[0]["last_provider"], "openai")
        self.assertNotIn("reason", suggestions[0])

    def test_security_job_stays_external_and_uses_infrastructure_as_nearest_hint(self):
        match = match_catalog_job(self.db, "网络安全工程师")

        self.assertEqual(match["catalog_status"], "external")
        self.assertEqual(match["nearest_catalog_job_name"], "DevOps 工程师")
        self.assertEqual(match["match_method"], "semantic_hint")

        refreshed = refresh_resume_direction_matches(
            self.db,
            [{"name": "网络安全工程师", "nearest_catalog_job_name": "数据工程师", "score": 75}],
        )
        self.assertEqual(refreshed[0]["nearest_catalog_job_name"], "DevOps 工程师")
        self.assertEqual(refreshed[0]["score"], 75)

    def test_admin_can_review_or_merge_an_external_job_suggestion(self):
        external = enrich_resume_directions(self.db, [{"name": "服务端研发工程师"}], "qwen")[0]
        suggestion_id = external["suggestion_id"]
        approved = review_job_suggestion(self.db, suggestion_id, "approved", "admin-1", "建议下一版新增")
        self.assertEqual(approved["review_status"], "approved")

        backend_job = match_catalog_job(self.db, "后端开发工程师")
        merged = merge_job_suggestion(self.db, suggestion_id, backend_job["catalog_job_id"], "admin-1")
        self.assertEqual(merged["review_status"], "merged")
        future_match = match_catalog_job(self.db, "服务端研发工程师")
        self.assertEqual(future_match["catalog_status"], "matched")
        self.assertEqual(future_match["match_method"], "alias")
        self.assertEqual(future_match["catalog_job_id"], backend_job["catalog_job_id"])

    def test_draft_entities_can_be_created_updated_disabled_and_deleted(self):
        version = create_version(self.db, {"code": "manual-v2", "name": "手工维护目录"}, "admin-1")
        college = create_catalog_entity(
            self.db,
            version["id"],
            "colleges",
            {"code": "ART", "name": "艺术学院", "sort_order": 10},
        )
        self.assertEqual(college["enabled"], 1)

        updated = update_catalog_entity(
            self.db,
            version["id"],
            "colleges",
            college["id"],
            {"name": "设计与艺术学院", "enabled": False, "sort_order": 20},
        )
        self.assertEqual(updated["name"], "设计与艺术学院")
        self.assertEqual(updated["enabled"], 0)

        self.assertTrue(delete_catalog_entity(self.db, version["id"], "colleges", college["id"]))
        self.assertFalse(delete_catalog_entity(self.db, version["id"], "colleges", college["id"]))
        with self.assertRaisesRegex(ValueError, "草稿"):
            create_catalog_entity(self.db, self.pilot_id, "colleges", {"code": "NEW", "name": "新学院"})

    def test_approved_suggestion_is_written_to_draft_with_ability_matrix(self):
        version = create_version(self.db, {"code": "school-v3", "name": "全校目录 V3"}, "admin-1")
        import_catalog_excel(self.db, version["id"], "catalog.xlsx", build_workbook(), "replace", "admin-1")
        draft_tree = catalog_tree(self.db, "school-v3", include_drafts=True)
        employment = draft_tree["colleges"][0]["majors"][0]["training_directions"][0]["employment_directions"][0]
        competency = draft_tree["competencies"][0]
        suggestion = enrich_resume_directions(self.db, [{"name": "增长产品经理"}], "qwen")[0]

        result = approve_job_suggestion_to_draft(
            self.db,
            suggestion["suggestion_id"],
            version["id"],
            {
                "employment_direction_id": employment["id"],
                "code": "GROWTH-PM",
                "competencies": [
                    {"competency_id": competency["id"], "required_level": 4, "weight": 5, "required": True}
                ],
            },
            "admin-1",
        )

        self.assertEqual(result["suggestion"]["review_status"], "approved")
        self.assertEqual(result["job_role"]["name"], "增长产品经理")
        self.assertEqual(result["competencies"][0]["required_level"], 4)
        approved = list_job_suggestions(self.db, "approved")
        self.assertEqual(approved[0]["draft_version_code"], "school-v3")
        self.assertEqual(approved[0]["draft_job_name"], "增长产品经理")


if __name__ == "__main__":
    unittest.main()
